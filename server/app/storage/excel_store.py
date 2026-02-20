"""Excel 存储层 - 开发阶段临时使用

数据存储结构：
- users.xlsx: 用户数据
- entities.xlsx: 客户/APP/模板数据（三个 sheet）
- experiments.xlsx: 实验数据（experiments, experiment_templates, experiment_groups）
"""

import threading
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
import json

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings


# 预设颜色列表
PRESET_COLORS = [
    "#FF6B6B", "#4ECDC4", "#45B7D1", "#96CEB4", "#FFEAA7",
    "#DDA0DD", "#98D8C8", "#F7DC6F", "#BB8FCE", "#85C1E9"
]


def get_color_for_experiment(experiment_id: int) -> str:
    """根据实验ID获取点缀色"""
    return PRESET_COLORS[experiment_id % len(PRESET_COLORS)]


class ExcelStore:
    """Excel 存储管理器"""

    _instance = None
    _lock = threading.Lock()

    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return

        self._initialized = True
        self.data_dir = settings.storage_path / "excel_data"
        self.data_dir.mkdir(parents=True, exist_ok=True)

        self._file_lock = threading.Lock()

        # 初始化文件
        self._init_files()

    def _init_files(self):
        """初始化 Excel 文件"""
        # 用户文件
        users_file = self.data_dir / "users.xlsx"
        if not users_file.exists():
            self._create_users_file(users_file)

        # 实体文件（客户/APP/模板）
        entities_file = self.data_dir / "entities.xlsx"
        if not entities_file.exists():
            self._create_entities_file(entities_file)

        # 实验文件
        experiments_file = self.data_dir / "experiments.xlsx"
        if not experiments_file.exists():
            self._create_experiments_file(experiments_file)

    def _create_users_file(self, filepath: Path):
        """创建用户文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "users"

        # 表头
        headers = ["id", "username", "password_hash", "display_name", "role",
                   "is_active", "must_change_password", "created_at", "created_by", "last_login_at"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 默认 root 用户 (密码: root123)
        from app.core.security import get_password_hash
        default_user = [
            1, "root", get_password_hash("root123"), "管理员", "root",
            True, True, datetime.now().isoformat(), None, None
        ]
        for col, value in enumerate(default_user, 1):
            ws.cell(row=2, column=col, value=value)

        wb.save(filepath)

    def _create_entities_file(self, filepath: Path):
        """创建实体文件"""
        wb = Workbook()

        # 客户表
        ws_customers = wb.active
        ws_customers.title = "customers"
        headers = ["id", "name", "contact", "description", "created_at"]
        for col, header in enumerate(headers, 1):
            ws_customers.cell(row=1, column=col, value=header)

        # 应用表
        ws_apps = wb.create_sheet("apps")
        headers = ["id", "customer_id", "name", "description", "created_at"]
        for col, header in enumerate(headers, 1):
            ws_apps.cell(row=1, column=col, value=header)

        # 模板表
        ws_templates = wb.create_sheet("templates")
        headers = ["id", "app_id", "name", "description", "created_at"]
        for col, header in enumerate(headers, 1):
            ws_templates.cell(row=1, column=col, value=header)

        wb.save(filepath)

    def _create_experiments_file(self, filepath: Path):
        """创建实验文件"""
        wb = Workbook()

        # 实验表
        ws_experiments = wb.active
        ws_experiments.title = "experiments"
        headers = ["id", "name", "status", "reference_type", "color",
                   "created_at", "created_by", "updated_at"]
        for col, header in enumerate(headers, 1):
            ws_experiments.cell(row=1, column=col, value=header)

        # 实验-模板关联表
        ws_links = wb.create_sheet("experiment_templates")
        headers = ["experiment_id", "template_id"]
        for col, header in enumerate(headers, 1):
            ws_links.cell(row=1, column=col, value=header)

        # 实验组表
        ws_groups = wb.create_sheet("experiment_groups")
        headers = ["id", "experiment_id", "name", "encoder_version", "transcode_params",
                   "input_url", "output_url", "status", "order_index", "created_at", "updated_at"]
        for col, header in enumerate(headers, 1):
            ws_groups.cell(row=1, column=col, value=header)

        # 客观指标表
        ws_metrics = wb.create_sheet("objective_metrics")
        headers = ["id", "group_id", "bitrate", "vmaf", "psnr", "ssim",
                   "machine_type", "concurrent_streams", "cpu_usage", "gpu_usage",
                   "detailed_report_url", "created_at", "updated_at"]
        for col, header in enumerate(headers, 1):
            ws_metrics.cell(row=1, column=col, value=header)

        wb.save(filepath)

    # ============ 通用方法 ============

    def _load_workbook(self, filename: str) -> Workbook:
        """加载工作簿"""
        filepath = self.data_dir / filename
        with self._file_lock:
            return load_workbook(filepath)

    def _save_workbook(self, wb: Workbook, filename: str):
        """保存工作簿"""
        filepath = self.data_dir / filename
        with self._file_lock:
            wb.save(filepath)

    def _get_next_id(self, ws: Worksheet) -> int:
        """获取下一个ID"""
        max_id = 0
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if value and isinstance(value, int):
                max_id = max(max_id, value)
        return max_id + 1

    def _row_to_dict(self, ws: Worksheet, row: int) -> Optional[Dict[str, Any]]:
        """行转字典"""
        if row > ws.max_row:
            return None

        values = []
        for col in range(1, ws.max_column + 1):
            values.append(ws.cell(row=row, column=col).value)

        if not values or values[0] is None:
            return None

        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        return dict(zip(headers, values))

    def _find_row_by_id(self, ws: Worksheet, id_value: int) -> int:
        """根据ID查找行号，返回0表示未找到"""
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == id_value:
                return row
        return 0

    # ============ 用户方法 ============

    def get_user_by_username(self, username: str) -> Optional[Dict]:
        """根据用户名获取用户"""
        wb = self._load_workbook("users.xlsx")
        ws = wb["users"]

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == username:
                return self._row_to_dict(ws, row)
        return None

    def get_user_by_id(self, user_id: int) -> Optional[Dict]:
        """根据ID获取用户"""
        wb = self._load_workbook("users.xlsx")
        ws = wb["users"]

        row = self._find_row_by_id(ws, user_id)
        if row:
            return self._row_to_dict(ws, row)
        return None

    def list_users(self) -> List[Dict]:
        """获取所有用户"""
        wb = self._load_workbook("users.xlsx")
        ws = wb["users"]

        users = []
        for row in range(2, ws.max_row + 1):
            user = self._row_to_dict(ws, row)
            if user:
                users.append(user)
        return users

    def create_user(self, username: str, password_hash: str, display_name: str,
                    role: str, created_by: int = None) -> Dict:
        """创建用户"""
        wb = self._load_workbook("users.xlsx")
        ws = wb["users"]

        new_id = self._get_next_id(ws)
        new_row = ws.max_row + 1

        ws.cell(row=new_row, column=1, value=new_id)
        ws.cell(row=new_row, column=2, value=username)
        ws.cell(row=new_row, column=3, value=password_hash)
        ws.cell(row=new_row, column=4, value=display_name)
        ws.cell(row=new_row, column=5, value=role)
        ws.cell(row=new_row, column=6, value=True)
        ws.cell(row=new_row, column=7, value=True)
        ws.cell(row=new_row, column=8, value=datetime.now().isoformat())
        ws.cell(row=new_row, column=9, value=created_by)
        ws.cell(row=new_row, column=10, value=None)

        self._save_workbook(wb, "users.xlsx")
        return self._row_to_dict(ws, new_row)

    def update_user(self, user_id: int, **kwargs) -> Optional[Dict]:
        """更新用户"""
        wb = self._load_workbook("users.xlsx")
        ws = wb["users"]

        row = self._find_row_by_id(ws, user_id)
        if not row:
            return None

        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        for key, value in kwargs.items():
            if key in headers:
                col = headers.index(key) + 1
                ws.cell(row=row, column=col, value=value)

        self._save_workbook(wb, "users.xlsx")
        return self._row_to_dict(ws, row)

    def delete_user(self, user_id: int) -> bool:
        """删除用户"""
        wb = self._load_workbook("users.xlsx")
        ws = wb["users"]

        row = self._find_row_by_id(ws, user_id)
        if not row:
            return False

        ws.delete_rows(row)
        self._save_workbook(wb, "users.xlsx")
        return True

    # ============ 客户方法 ============

    def list_customers(self) -> List[Dict]:
        """获取所有客户"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["customers"]

        customers = []
        for row in range(2, ws.max_row + 1):
            customer = self._row_to_dict(ws, row)
            if customer:
                customers.append(customer)
        return customers

    def get_customer_by_id(self, customer_id: int) -> Optional[Dict]:
        """根据ID获取客户"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["customers"]

        row = self._find_row_by_id(ws, customer_id)
        if row:
            return self._row_to_dict(ws, row)
        return None

    def create_customer(self, name: str, contact: str = None, description: str = None) -> Dict:
        """创建客户"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["customers"]

        new_id = self._get_next_id(ws)
        new_row = ws.max_row + 1

        ws.cell(row=new_row, column=1, value=new_id)
        ws.cell(row=new_row, column=2, value=name)
        ws.cell(row=new_row, column=3, value=contact)
        ws.cell(row=new_row, column=4, value=description)
        ws.cell(row=new_row, column=5, value=datetime.now().isoformat())

        self._save_workbook(wb, "entities.xlsx")
        return self._row_to_dict(ws, new_row)

    def update_customer(self, customer_id: int, **kwargs) -> Optional[Dict]:
        """更新客户"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["customers"]

        row = self._find_row_by_id(ws, customer_id)
        if not row:
            return None

        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        for key, value in kwargs.items():
            if key in headers:
                col = headers.index(key) + 1
                ws.cell(row=row, column=col, value=value)

        self._save_workbook(wb, "entities.xlsx")
        return self._row_to_dict(ws, row)

    def delete_customer(self, customer_id: int) -> bool:
        """删除客户"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["customers"]

        row = self._find_row_by_id(ws, customer_id)
        if not row:
            return False

        ws.delete_rows(row)
        self._save_workbook(wb, "entities.xlsx")
        return True

    # ============ 应用方法 ============

    def list_apps(self, customer_id: int = None) -> List[Dict]:
        """获取应用列表"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["apps"]

        apps = []
        for row in range(2, ws.max_row + 1):
            app = self._row_to_dict(ws, row)
            if app and (customer_id is None or app.get("customer_id") == customer_id):
                apps.append(app)
        return apps

    def get_app_by_id(self, app_id: int) -> Optional[Dict]:
        """根据ID获取应用"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["apps"]

        row = self._find_row_by_id(ws, app_id)
        if row:
            return self._row_to_dict(ws, row)
        return None

    def create_app(self, customer_id: int, name: str, description: str = None) -> Dict:
        """创建应用"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["apps"]

        new_id = self._get_next_id(ws)
        new_row = ws.max_row + 1

        ws.cell(row=new_row, column=1, value=new_id)
        ws.cell(row=new_row, column=2, value=customer_id)
        ws.cell(row=new_row, column=3, value=name)
        ws.cell(row=new_row, column=4, value=description)
        ws.cell(row=new_row, column=5, value=datetime.now().isoformat())

        self._save_workbook(wb, "entities.xlsx")
        return self._row_to_dict(ws, new_row)

    def update_app(self, app_id: int, **kwargs) -> Optional[Dict]:
        """更新应用"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["apps"]

        row = self._find_row_by_id(ws, app_id)
        if not row:
            return None

        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        for key, value in kwargs.items():
            if key in headers:
                col = headers.index(key) + 1
                ws.cell(row=row, column=col, value=value)

        self._save_workbook(wb, "entities.xlsx")
        return self._row_to_dict(ws, row)

    def delete_app(self, app_id: int) -> bool:
        """删除应用"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["apps"]

        row = self._find_row_by_id(ws, app_id)
        if not row:
            return False

        ws.delete_rows(row)
        self._save_workbook(wb, "entities.xlsx")
        return True

    # ============ 模板方法 ============

    def list_templates(self, app_id: int = None) -> List[Dict]:
        """获取模板列表"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["templates"]

        templates = []
        for row in range(2, ws.max_row + 1):
            template = self._row_to_dict(ws, row)
            if template and (app_id is None or template.get("app_id") == app_id):
                templates.append(template)
        return templates

    def get_template_by_id(self, template_id: int) -> Optional[Dict]:
        """根据ID获取模板"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["templates"]

        row = self._find_row_by_id(ws, template_id)
        if row:
            return self._row_to_dict(ws, row)
        return None

    def create_template(self, app_id: int, name: str, description: str = None) -> Dict:
        """创建模板"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["templates"]

        new_id = self._get_next_id(ws)
        new_row = ws.max_row + 1

        ws.cell(row=new_row, column=1, value=new_id)
        ws.cell(row=new_row, column=2, value=app_id)
        ws.cell(row=new_row, column=3, value=name)
        ws.cell(row=new_row, column=4, value=description)
        ws.cell(row=new_row, column=5, value=datetime.now().isoformat())

        self._save_workbook(wb, "entities.xlsx")
        return self._row_to_dict(ws, new_row)

    def update_template(self, template_id: int, **kwargs) -> Optional[Dict]:
        """更新模板"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["templates"]

        row = self._find_row_by_id(ws, template_id)
        if not row:
            return None

        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        for key, value in kwargs.items():
            if key in headers:
                col = headers.index(key) + 1
                ws.cell(row=row, column=col, value=value)

        self._save_workbook(wb, "entities.xlsx")
        return self._row_to_dict(ws, row)

    def delete_template(self, template_id: int) -> bool:
        """删除模板"""
        wb = self._load_workbook("entities.xlsx")
        ws = wb["templates"]

        row = self._find_row_by_id(ws, template_id)
        if not row:
            return False

        ws.delete_rows(row)
        self._save_workbook(wb, "entities.xlsx")
        return True

    # ============ 实验方法 ============

    def list_experiments(self, template_id: int = None, status: str = None,
                         page: int = 1, page_size: int = 20) -> tuple[List[Dict], int]:
        """获取实验列表"""
        wb = self._load_workbook("experiments.xlsx")
        ws = wb["experiments"]
        ws_links = wb["experiment_templates"]

        # 获取所有实验
        all_experiments = []
        for row in range(2, ws.max_row + 1):
            exp = self._row_to_dict(ws, row)
            if exp:
                all_experiments.append(exp)

        # 如果指定了 template_id，需要过滤
        if template_id:
            # 获取关联的实验ID
            linked_exp_ids = set()
            for row in range(2, ws_links.max_row + 1):
                if ws_links.cell(row=row, column=2).value == template_id:
                    linked_exp_ids.add(ws_links.cell(row=row, column=1).value)

            all_experiments = [e for e in all_experiments if e["id"] in linked_exp_ids]

        # 如果指定了状态
        if status:
            all_experiments = [e for e in all_experiments if e.get("status") == status]

        total = len(all_experiments)

        # 分页
        start = (page - 1) * page_size
        end = start + page_size
        experiments = all_experiments[start:end]

        return experiments, total

    def get_experiment_by_id(self, experiment_id: int) -> Optional[Dict]:
        """根据ID获取实验"""
        wb = self._load_workbook("experiments.xlsx")
        ws = wb["experiments"]

        row = self._find_row_by_id(ws, experiment_id)
        if row:
            return self._row_to_dict(ws, row)
        return None

    def get_experiment_template_ids(self, experiment_id: int) -> List[int]:
        """获取实验关联的模板ID列表"""
        wb = self._load_workbook("experiments.xlsx")
        ws = wb["experiment_templates"]

        template_ids = []
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == experiment_id:
                template_ids.append(ws.cell(row=row, column=2).value)
        return template_ids

    def create_experiment(self, name: str, template_ids: List[int], created_by: int,
                          status: str = "draft", reference_type: str = "new") -> Dict:
        """创建实验"""
        wb = self._load_workbook("experiments.xlsx")
        ws = wb["experiments"]
        ws_links = wb["experiment_templates"]

        new_id = self._get_next_id(ws)
        color = get_color_for_experiment(new_id)
        new_row = ws.max_row + 1

        ws.cell(row=new_row, column=1, value=new_id)
        ws.cell(row=new_row, column=2, value=name)
        ws.cell(row=new_row, column=3, value=status)
        ws.cell(row=new_row, column=4, value=reference_type)
        ws.cell(row=new_row, column=5, value=color)
        ws.cell(row=new_row, column=6, value=datetime.now().isoformat())
        ws.cell(row=new_row, column=7, value=created_by)
        ws.cell(row=new_row, column=8, value=None)

        # 添加模板关联
        for template_id in template_ids:
            link_row = ws_links.max_row + 1
            ws_links.cell(row=link_row, column=1, value=new_id)
            ws_links.cell(row=link_row, column=2, value=template_id)

        self._save_workbook(wb, "experiments.xlsx")
        return self._row_to_dict(ws, new_row)

    def update_experiment(self, experiment_id: int, **kwargs) -> Optional[Dict]:
        """更新实验"""
        wb = self._load_workbook("experiments.xlsx")
        ws = wb["experiments"]

        row = self._find_row_by_id(ws, experiment_id)
        if not row:
            return None

        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        for key, value in kwargs.items():
            if key in headers:
                col = headers.index(key) + 1
                ws.cell(row=row, column=col, value=value)

        # 更新 updated_at
        updated_at_col = headers.index("updated_at") + 1
        ws.cell(row=row, column=updated_at_col, value=datetime.now().isoformat())

        self._save_workbook(wb, "experiments.xlsx")
        return self._row_to_dict(ws, row)

    def delete_experiment(self, experiment_id: int) -> bool:
        """删除实验"""
        wb = self._load_workbook("experiments.xlsx")
        ws = wb["experiments"]
        ws_links = wb["experiment_templates"]

        row = self._find_row_by_id(ws, experiment_id)
        if not row:
            return False

        # 删除实验
        ws.delete_rows(row)

        # 删除关联
        rows_to_delete = []
        for row in range(2, ws_links.max_row + 1):
            if ws_links.cell(row=row, column=1).value == experiment_id:
                rows_to_delete.append(row)

        for row in reversed(rows_to_delete):
            ws_links.delete_rows(row)

        self._save_workbook(wb, "experiments.xlsx")
        return True

    def link_experiment_templates(self, experiment_id: int, template_ids: List[int]):
        """关联实验和模板"""
        wb = self._load_workbook("experiments.xlsx")
        ws_links = wb["experiment_templates"]

        existing = set()
        for row in range(2, ws_links.max_row + 1):
            if ws_links.cell(row=row, column=1).value == experiment_id:
                existing.add(ws_links.cell(row=row, column=2).value)

        for template_id in template_ids:
            if template_id not in existing:
                new_row = ws_links.max_row + 1
                ws_links.cell(row=new_row, column=1, value=experiment_id)
                ws_links.cell(row=new_row, column=2, value=template_id)

        self._save_workbook(wb, "experiments.xlsx")

    def unlink_experiment_template(self, experiment_id: int, template_id: int):
        """解除实验和模板的关联"""
        wb = self._load_workbook("experiments.xlsx")
        ws_links = wb["experiment_templates"]

        rows_to_delete = []
        for row in range(2, ws_links.max_row + 1):
            if (ws_links.cell(row=row, column=1).value == experiment_id and
                ws_links.cell(row=row, column=2).value == template_id):
                rows_to_delete.append(row)

        for row in reversed(rows_to_delete):
            ws_links.delete_rows(row)

        self._save_workbook(wb, "experiments.xlsx")

    # ============ 矩阵数据方法 ============

    def get_matrix_data(self) -> tuple[List[Dict], List[Dict]]:
        """获取矩阵数据"""
        wb_entities = self._load_workbook("entities.xlsx")
        wb_experiments = self._load_workbook("experiments.xlsx")

        # 获取所有模板
        ws_templates = wb_entities["templates"]
        ws_apps = wb_entities["apps"]
        ws_customers = wb_entities["customers"]
        ws_experiments = wb_experiments["experiments"]
        ws_links = wb_experiments["experiment_templates"]

        # 构建索引
        apps = {r["id"]: r for r in [self._row_to_dict(ws_apps, row) for row in range(2, ws_apps.max_row + 1)] if r}
        customers = {r["id"]: r for r in [self._row_to_dict(ws_customers, row) for row in range(2, ws_customers.max_row + 1)] if r}

        # 获取所有实验
        experiments = []
        for row in range(2, ws_experiments.max_row + 1):
            exp = self._row_to_dict(ws_experiments, row)
            if exp:
                if not exp.get("color"):
                    exp["color"] = get_color_for_experiment(exp["id"])
                experiments.append(exp)

        # 构建实验-模板关联索引
        exp_template_links = {}  # template_id -> [experiment_ids]
        for row in range(2, ws_links.max_row + 1):
            exp_id = ws_links.cell(row=row, column=1).value
            template_id = ws_links.cell(row=row, column=2).value
            if template_id not in exp_template_links:
                exp_template_links[template_id] = []
            exp_template_links[template_id].append(exp_id)

        # 构建实验索引
        exp_index = {e["id"]: e for e in experiments}

        # 构建矩阵行
        rows = []
        for row in range(2, ws_templates.max_row + 1):
            template = self._row_to_dict(ws_templates, row)
            if not template:
                continue

            app = apps.get(template.get("app_id"))
            if not app:
                continue

            customer = customers.get(app.get("customer_id"))
            if not customer:
                continue

            # 获取该模板关联的实验
            linked_exp_ids = exp_template_links.get(template["id"], [])
            template_experiments = {}
            for exp_id in linked_exp_ids:
                if exp_id in exp_index:
                    exp = exp_index[exp_id]
                    template_experiments[exp_id] = {
                        "id": exp["id"],
                        "name": exp["name"],
                        "status": exp["status"],
                        "color": exp.get("color", get_color_for_experiment(exp["id"]))
                    }

            rows.append({
                "customer_id": customer["id"],
                "customer_name": customer["name"],
                "app_id": app["id"],
                "app_name": app["name"],
                "template_id": template["id"],
                "template_name": template["name"],
                "experiments": template_experiments
            })

        return rows, experiments


# 全局实例
excel_store = ExcelStore()
